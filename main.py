##############################################
#  Autor: Theodore Noubissi Kala
#
#  Aufgabe: Grafische Oberfläsche für Zahnartz
##############################################

import tkinter as tk
from tkinter import colorchooser
import re
import pandas as pd
import openpyxl
from tkinter import filedialog
from tkinter import ttk
from datetime import datetime
from tkinter import messagebox
from PIL import ImageTk, Image, Image
from tkinter import Label, PhotoImage
from tkinter.ttk import Combobox
from tkinter.font import Font
from tkcalendar import Calendar
from openpyxl import load_workbook

#database für Ärzte (beim Login)
db = pd.read_excel("Database/Patienten_Zahnärzte_Kosten.xlsx", sheet_name="Zahnärzte", header=None)
db2 = pd.read_excel("Database/Patienten_Zahnärzte_Kosten.xlsx", sheet_name="Stamm-Patienten2", header=None)
db3 = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="Stamm-Patienten2")


class LoginWindow(tk.Tk):
    def __init__(self): # Funktion für die LoginWindow-Klasse initialisieren
        super().__init__()
        self.title("Anmeldung")
        self.config(background="#3B6064")

        # Login Seite
        login_Frame = tk.Frame(background="#5B949A", width=450, height= 850)
        login_Frame.pack(pady=60, padx=60)

        Login_label = tk.Label(login_Frame, background="#5B949A", text="Log in", fg="white", font=("Arial", 23))
        Login_label.place(x=170, y=35)


        # Benutzername-Etikett und Eingabefeld
        username_label = tk.Label(login_Frame, text="Name", background="#5B949A", fg="white", font=("Arial", 15))
        username_label.place(x= 70, y=120, height=36)
        password_label = tk.Label(login_Frame, text="Passwort", background="#5B949A", fg="white", font=("Arial", 15))
        password_label.place(x= 70, y=180, height=36)


        self.username_entry = tk.Entry(login_Frame, background="white", fg="black", font=("Arial", 13))
        self.username_entry.place(x= 190, y=120, height=36, width=180)
        self.password_entry = tk.Entry(login_Frame, background="white", fg="black", show="*", font=("Arial", 13))
        self.password_entry.place(x= 190, y=180, height=36, width=182) 

        # Anmeldebutton und Registrierungsbutton
        login_button = tk.Button(login_Frame, text="Login", width= 41, height= 2, background="#3B6064", fg="white", command=self.login)   # TODO: command=self.login implementieren 
        login_button.place(x=70, y=260)        
        arzt_regist_button = tk.Button(login_Frame, text="Patient sign Up", width= 17, height= 2, background="#3B6064", fg="white", command=self.open_new_window_Patient_regist)   # TODO: command=self.open_new_window_Patient_regist implementieren 
        arzt_regist_button.place(x=70, y=320)        
        patient_regist_button = tk.Button(login_Frame, text="Artz sign Up", width= 17, height= 2, background="#3B6064", fg="white", command = self.open_new_window_Arzt_regist)   # TODO: command=self.open_new_window_Arzt_regist implementieren 
        patient_regist_button.place(x=237, y=320)   

        # Attribut pour stocker le nom d'utilisateur actuel
        self.current_user = None     

    def login(self):
        password = None

        # Ärzten Database
        full_names = db.iloc[1:80, 0].values.tolist()
        last_names = [name.split()[-1] for name in full_names]
        passwords = db.iloc[1:80, 1].values.tolist()

        #Patienten
        full_names2 = db3.iloc[1:61, 0].values.tolist()
        patient_names = [name.split()[-1] for name in full_names2]
        patient_passwords = db3.iloc[1:61, 1].values.tolist()

        global username
        # global KrankenkasseArt
        username = self.username_entry.get()
        password = self.password_entry.get()

        if username in last_names and password in passwords:
            self.current_user = username
            # Vous pouvez maintenant utiliser la variable krankenkasse_art comme nécessaire
            self.open_dentist_view(username)
        elif username == "a" and password == "a":
            self.open_dentist_view(username)
        elif username == "p" and password == "p":
            self.open_Patient_View(username)
        elif username in patient_names and password in patient_passwords:

            # Create a list of possible formats for the username prefix
            possible_prefixes = ["Herr", "Frau"]

            # Variable to store the prefix
            prefix = None

            # Check each possible format
            for possible_prefix in possible_prefixes:
                formatted_username = f"{possible_prefix} {username}"
                if db3['Patient'].str.contains(formatted_username, case=False).any():
                    prefix = possible_prefix
                    break

            possible_usernames = ["Herr " + username, "Frau " + username]
            matching_rows = db3['Patient'].str.contains('|'.join(possible_usernames), case=False)
            # krankenkasse_art = db3.loc[db3.iloc[:, 0] == username, db3.columns[2]].values[0]
            krankenkasse_art = db3.loc[matching_rows, 'Krankenkassenart'].values[0]
            self.open_Patient_View(username, krankenkasse_art, prefix) 
        else:
            messagebox.showerror("Fehler", "Ungultige Anmeldeinformationen")

    def open_dentist_view(self, last_names):
        self.withdraw()
        # self.iconify()
        dentist_view = DentistView(self, last_names)
        dentist_view.mainloop()

    def open_Patient_View(self, patient_names, krankenkasse_art, prefix):
        self.withdraw()
        patient_View = PatientView(self, patient_names, krankenkasse_art, prefix)
        patient_View.mainloop()

    def open_new_window_Arzt_regist(self):
        self.withdraw()  

        new_window = tk.Toplevel()
        new_window.title("Arzt Registrierung")
        new_window.geometry("700x450")
        new_window.config(background="#2A324B")

        # Variables
        self.name_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.problematik_var = tk.StringVar()
        self.dentist_var = tk.StringVar()
        self.teeth_var = tk.StringVar()
  

        name_Label = tk.Label(new_window, text="Name:", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        name_Label.place(rely=0.2, relx=0.09)
        name_Label_Entry = tk.Entry(new_window, textvariable=self.name_var)
        name_Label_Entry.place(rely=0.2, relx=0.29)

        passwort_Label = tk.Label(new_window, text="passwort:", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        passwort_Label.place(rely=0.3, relx=0.09)
        passwort_Label_Entry = tk.Entry(new_window, textvariable=self.password_var, show ='*')
        passwort_Label_Entry.place(rely=0.30, relx=0.29)
        
        Krankenkassenart_Label = tk.Label(new_window, text="Krankenkassenart:", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        Krankenkassenart_Label.place(rely=0.4, relx=0.09)
        krankenkassenarten = ["gesetzlich", "freiwillig gesetzlich", "privat"]
        self.krankenk_var = tk.StringVar(self)
        self.combobox2 = Combobox(new_window, values=krankenkassenarten, textvariable=self.krankenk_var, width=24)  # Bind the selection event
        self.combobox2.current(0)  # Set the default selection
        self.combobox2.place(rely=0.4, relx=0.29)

        # register_button = tk.Button(new_window, text="Register", font=("Arial", 11, "bold"), command=self.save_data)
        self.register_button = tk.Button(new_window, text="Registrieren", font=("Arial", 11, "bold"), command=self.save_data)
        self.register_button.place(rely=0.55, relx=0.19)

        self.Auslogen_button = tk.Button(new_window, text="Zurück",command=self.logout, width=25, bg="#DE5466", fg="white", font=("Arial", 9, "bold"))
        self.Auslogen_button.pack(side="bottom", pady=7)

        name_Label_Entry.delete(0, tk.END)
        passwort_Label_Entry.delete(0, tk.END)
        self.combobox2.delete(0, tk.END)
    	
        # new_window.protocol("WM_DELETE_WINDOW", self.on_new_window_close)
    
    def open_new_window_Patient_regist(self):
        # self.withdraw()  

        new_window = tk.Toplevel()
        new_window.title("Patient Registrierung")
        new_window.geometry("700x450")
        new_window.config(background="#2A324B")


        # Variables
        self.name_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.problematik_var = tk.StringVar()
        self.dentist_var = tk.StringVar()
        self.teeth_var = tk.StringVar()
  

        name_Label = tk.Label(new_window, text="Name:", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        name_Label.place(rely=0.1, relx=0.05)
        name_Label_Entry = tk.Entry(new_window, textvariable=self.name_var, width=27)
        name_Label_Entry.place(rely=0.1, relx=0.25)

        passwort_Label = tk.Label(new_window, text="passwort:", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        passwort_Label.place(rely=0.2, relx=0.05)
        passwort_Label_Entry = tk.Entry(new_window, show ='*', textvariable=self.password_var, width=27)
        passwort_Label_Entry.place(rely=0.20, relx=0.25)
        
        Krankenkassenart_Label = tk.Label(new_window, text="Krankenkassenart:", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        Krankenkassenart_Label.place(rely=0.3, relx=0.05)
        krankenkassenarten = ["gesetzlich", "freiwillig gesetzlich", "privat"]
        self.krankenk_var = tk.StringVar(self)
        self.combobox2 = Combobox(new_window, values=krankenkassenarten, textvariable=self.krankenk_var, width=24)  # Bind the selection event
        self.combobox2.current(0)  # Set the default selection
        self.combobox2.place(rely=0.3, relx=0.25)

        Problematik_label = tk.Label(new_window, text="Problematik", font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        Problematik_label.place(rely=0.4, relx=0.05)
        Problem_options = ["Karies klein", "Karies groß", "Teilkrone", "krone", "Wurzelbehandlung"]
        self.combobox1 = Combobox(new_window, textvariable=self.problematik_var, values=Problem_options, width=24)
        self.problematik_var.set(Problem_options[0])  # Set the default selection
        self.combobox1.place(rely=0.4, relx=0.25)   

        self.Zahnanzahl_label = tk.Label(new_window, text="Zahnanzahl",  font=("Arial", 10, "bold"), fg="white", background="#2A324B")
        self.Zahnanzahl_label.place(rely=0.5, relx=0.05)
        self.Zahnanzahl_entry = tk.Entry(new_window, background="#f1f1f1", textvariable=self.teeth_var, width=27)
        self.Zahnanzahl_entry.place(rely=0.5, relx=0.25) 

        # register_button = tk.Button(new_window, text="Register", font=("Arial", 11, "bold"), command=self.save_data)
        self.register_button = tk.Button(new_window, text="Registrieren", font=("Arial", 11, "bold"), command=self.save_data2)
        self.register_button.place(rely=0.65, relx=0.16)

        self.Auslogen_button = tk.Button(new_window, text="Zurück",command=self.logout, width=25, bg="#DE5466", fg="white", font=("Arial", 9, "bold"))
        self.Auslogen_button.pack(side="bottom", pady=7)

        name_Label_Entry.delete(0, tk.END)
        passwort_Label_Entry.delete(0, tk.END)
        self.combobox2.delete(0, tk.END)
        self.combobox1.delete(0, tk.END)
        self.Zahnanzahl_entry.delete(0, tk.END)

        new_window.protocol("WM_DELETE_WINDOW", self.on_new_window_close)

    def save_data(self):
        # Récupérer les données saisies
        name = self.name_var.get()
        password = self.password_var.get()
        krankenkassenart = self.krankenk_var.get()

        # Charger le fichier Excel existant
        workbook = openpyxl.load_workbook('Database/Patienten_Zahnärzte_Kosten.xlsx')
        # Sélectionner la feuille de calcul spécifiée
        sheet = workbook['Zahnärzte']
        # Trouver la première ligne vide à partir de la ligne 4 (lignes 1 à 3 sont des en-têtes)
        row = 3
        while sheet.cell(row=row, column=2).value:
            row += 1
        # Enregistrer les données dans les colonnes spécifiées
        sheet.cell(row=row, column=2).value = name
        sheet.cell(row=row, column=3).value = password
        sheet.cell(row=row, column=4).value = krankenkassenart

        # Sauvegarder les modifications dans le fichier Excel
        workbook.save('Database/Patienten_Zahnärzte_Kosten.xlsx')

        # Afficher un message de succès
        tk.messagebox.showinfo("Success", "Registrierung Erfolgreich")

    def is_name_exists(self, name, sheet):
        # Vérifie si le nom existe déjà dans la colonne 3 (colonne des noms)
        for row in sheet.iter_rows(min_row=4, min_col=3, max_col=3):
            if row[0].value == name:
                return True
        return False

    def generate_unique_name(self, name, sheet):
        # Wenn ein Patient denselben Namen hat, generieren Sie einen eindeutigen Namen mit einer numerischen Erweiterung
        name_count = 0
        modified_name = name
        while self.is_name_exists(modified_name, sheet):
            name_count += 1
            modified_name = f"{name}_{name_count}"
        return modified_name

    def save_data2(self):
        # Daten aus den Eingabefeldern abrufen
        name = self.name_var.get()
        password = self.password_var.get()
        krankenkassenart = self.krankenk_var.get()
        zahn_probl = self.problematik_var.get()
        zahn_anzahl = self.teeth_var.get()

        # Vorhandenes Excel-Datei laden
        workbook = openpyxl.load_workbook('Database/Patienten_Zahnärzte_Kosten.xlsx')
        # Gewünschtes Tabellenblatt auswählen
        sheet = workbook['Stamm-Patienten2']

        # Überprüfen, ob der Name bereits in den Daten vorhanden ist
        if self.is_name_exists(name, sheet):
            # Wenn der Name bereits vorhanden ist, fragen Sie den Benutzer, ob er sich bereits registriert hat.
            response = tk.messagebox.askquestion("Frage", f"Es existiert bereits ein Eintrag für '{name}'. Sind Sie bereits registriert?")
            if response == 'yes':
                # Wenn der Benutzer bereits registriert ist, aktualisieren Sie das Passwort und die dentalen Probleme.
                for row in sheet.iter_rows(min_row=4, min_col=3, max_col=3):
                    if row[0].value == name:
                        row_index = row[0].row
                        # Aktualisieren Sie das Passwort und die anderen Spaltenwerte (z.B., Krankenkasse, etc.) nach Bedarf
                        sheet.cell(row=row_index, column=4).value = password
                        # Fügen Sie neue dental Probleme hinzu
                        current_problematik = sheet.cell(row=row_index, column=6).value
                        new_problematik = f"{current_problematik}, {zahn_probl}"
                        sheet.cell(row=row_index, column=6).value = new_problematik
                        break
            else:
                # Wenn der Benutzer nicht bereits registriert ist, können Sie hier entsprechende Aktionen durchführen.
                pass
        else:
            # Wenn der Name nicht vorhanden ist, generieren Sie einen eindeutigen Namen
            unique_name = self.generate_unique_name(name, sheet)

            # Finden Sie die erste leere Zeile ab Zeile 4 (Zeilen 1 bis 3 sind Überschriften).
            row = 4
            while sheet.cell(row=row, column=3).value:
                row += 1
            # Daten in die entsprechenden Spalten eintragen
            sheet.cell(row=row, column=1).value = unique_name
            sheet.cell(row=row, column=2).value = password
            sheet.cell(row=row, column=3).value = krankenkassenart
            sheet.cell(row=row, column=4).value = zahn_probl
            sheet.cell(row=row, column=5).value = zahn_anzahl

        # Sauvegarder les modifications dans le fichier Excel
        workbook.save('Database/Patienten_Zahnärzte_Kosten.xlsx')

        # Afficher un message de succès
        tk.messagebox.showinfo("Success", "Registrierung Erfolgreich")
    
    def logout(self):
        self.withdraw()
        self.deiconify()
        # self.master.deiconify()
        # self.master.username_entry.delete(0, tk.END)
        # self.master.password_entry.delete(0, tk.END)

    def get_password(self):
        return self.password_entry.get()

    def on_new_window_close(self):
        self.deiconify()  # Réaffiche la première fenêtre lorsque la deuxième fenêtre est fermée
        self.destroy()

    def open_change_password_window(self):
        change_password_window = ChangePassworWindow(self)
        change_password_window.title("Password ändern")

class PatientView(tk.Toplevel):
    def __init__(self, root, patient_names, krankenkasse_art, prefix):
        super().__init__()
        self.geometry("1300x568")
        self.minsize(1300, 568)
        self.maxsize(1300, 568)
        self.title("Patientenansicht")
        self.config(background="#3B6064")
        
        self.patient_names = patient_names
        self.krankenkasse_art = krankenkasse_art
        self.prefix = prefix

        # Simulierte Daten - Zahnärzte und ihre Termine
        self.zahnaerzte = {
            "Herr Dr. Kraft": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
            "Herr Dr. Hausmann": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
            "Frau Dr. Winkel": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
            "Herr Dr. Huber": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
            "Frau Dr. Wurzel": ["2023-11-10 11:00", "2023-11-10 15:30", "2023-11-10 16:00"]
        }


        self.root = root  # Conservez une référence à la fenêtre principale
        self.current_user = username

        # Hauptframe für das Layout der Patientenansicht
        self.main_frame = tk.Frame(self, width=240, height=580)
        self.main_frame.pack(side="left")

        # Linker Frame für die Navigation oder zusätzliche Informationen
        self.left_frame = tk.Frame(self.main_frame, width=240, height=580, bg="#5B949A")
        self.left_frame.pack(side="left")
        self.left_frame.pack_propagate(False)

        self.info_Frame = tk.LabelFrame(self.main_frame, width=238, height=305, background="white", text="")
        self.info_Frame.place(rely=0.003, relx=0.005)
        
        self.info_frame_unterbox = tk.Frame(self.info_Frame, width=230, height=150, background="white")
        self.info_frame_unterbox.place(rely=0.66, relx=0)

        self.image_frame = tk.Frame(self.info_Frame, width=180, height=180)
        self.image_frame.place(rely=0.04, relx=0.11)
        default_image = Image.open("sign-out-alt.png")  # Remplacez par votre image par défaut
        default_image = default_image.resize((180, 180), Image.LANCZOS)
        default_photo = ImageTk.PhotoImage(default_image)
        label = tk.Label(self.image_frame, image=default_photo, anchor="center")
        label.pack() 

        # Info Frame
        self.Name_Label = tk.Label(self.info_frame_unterbox ,text="Titel: " + prefix, font=("Arial", 11, "bold"), background="white")
        self.Name_Label.place(rely=0.08, relx=0.11)

        self.Name_Label = tk.Label(self.info_frame_unterbox ,text="Name: " + patient_names, font=("Arial", 11, "bold"), background="white")
        self.Name_Label.place(rely=0.25, relx=0.11)

        self.Krankenkass_Label = tk.Label(self.info_frame_unterbox ,text="Krankenkasse: " + self.krankenkasse_art, font=("Arial", 11, "bold"), background="white")
        self.Krankenkass_Label.place(rely=0.45, relx=0.11)

        #Button
        self.bg_button = tk.Button(self.left_frame, text="Background Ändern", width=25, bg="#2A324B", fg="white", font=("Arial", 9, "bold"))
        self.bg_button.pack(side="bottom", pady=13)

        self.Auslogen_button = tk.Button(self.left_frame, text="Auslogen",command=self.logout, width=25, bg="#DE5466", fg="white", font=("Arial", 9, "bold"))
        self.Auslogen_button.pack(side="bottom", pady=7)

        self.einstellung_button = tk.Button(self.left_frame, text="Passwort Ändern", width=25, bg="#2A324B", fg="white", font=("Arial", 9, "bold"), command=self.open_change_password_window) # TODO: , command=open_change_password_window
        self.einstellung_button.pack(side="bottom", pady=10)

        self.einstellung_button = tk.Button(self.left_frame, text="View Aktualisieren", width=25, bg="#2A324B", fg="white", font=("Arial", 9, "bold"), command=self.viewAktualisieren) # TODO: , command=open_change_password_window
        self.einstellung_button.pack(side="bottom", pady=10)
       
        self.untersten_Frame = tk.Frame(self, width=1058, height=30, background="#5B949A")
        self.untersten_Frame.pack(side="bottom", fill="y")

        self.firma_name_label = tk.Label(self.untersten_Frame, text="Zahnärzte am Park GmbH", background="#5B949A", fg="white", font=("Arial", 11, "bold"))
        # firma_name_label.pack(side="left", fill="x")
        self.firma_name_label.place(x=15, y=5)
        self.animate_label()  # Animation starten

        self.current_time_label = tk.Label(self.untersten_Frame, text="", font=("Arial", 11, "bold"), fg="white", bg="#5B949A")
        self.current_time_label.place(x=900, y=5)

        #--------Auswahl Frame oben---------
        auswahl_Frame1 = tk.LabelFrame(self, text="", width=1055, height=304, background="white")
        auswahl_Frame1.place(rely=0.003, relx=0.186)


        # Variables
        self.teeth_var = tk.StringVar()
        self.fill_material_var = tk.StringVar()
        self.cost_var = tk.StringVar()
        self.dentist_var = tk.StringVar()
        self.problematik_var = tk.StringVar()

        # Create labels and entries for teeth and filling selection
        self.DentaleProblem = tk.Label(auswahl_Frame1, text="Problematik:", background="white", font=("Arial", 10, "bold"))
        self.DentaleProblem.place(rely=0.05, relx=0.01)

        Problem_options = ["Karies klein", "Karies groß", "Teilkrone", "krone", "Wurzelbehandlung"]        
        self.combobox1 = Combobox(auswahl_Frame1, textvariable=self.problematik_var, values=Problem_options)
        self.problematik_var.set(Problem_options[0])  # Set the default selection
        self.combobox1.place(rely=0.05, relx=0.15) 

        self.teeth_label = tk.Label(auswahl_Frame1, text="Zahnanzahl:", background="white", font=("Arial", 10, "bold"))
        self.teeth_label.place(rely=0.243, relx=0.01)
        self.teeth_entry = tk.Entry(auswahl_Frame1, background="#f1f1f1",width=23 ,textvariable=self.teeth_var)
        self.teeth_entry.place(rely=0.243, relx=0.15)

        self.filling_label = tk.Label(auswahl_Frame1, text="Füllmaterial:", background="white", font=("Arial", 10, "bold"))
        self.filling_label.place(rely=0.45, relx=0.01)
        filling_options = ["normal", "höherwertig", "höchstwertig"]
        self.combobox1 = Combobox(auswahl_Frame1, textvariable=self.fill_material_var, values=filling_options)
        self.fill_material_var.set(filling_options[0])  # Set the default selection
        self.combobox1.place(rely=0.45, relx=0.15)

        self.behandlung_Zeit_Label = tk.Label(auswahl_Frame1,text="Behandlungszeit", background="white", font=("Arial", 10, "bold"))
        self.behandlung_Zeit_Label.place(rely=0.65, relx=0.01)
        self.anzeige_Zeit_Label = tk.Label(auswahl_Frame1,text="", height=2, width=10, background="#B2CFD2", font=("Arial", 10, "bold"))
        self.anzeige_Zeit_Label.place(rely=0.62, relx=0.15)

        self.Krankenk_label = tk.Label(auswahl_Frame1, text="Krankenkassenart:", background="white", font=("Arial", 10, "bold"))
        self.Krankenk_label.place(rely=0.05, relx=0.36)
        # Krankenkassenart ComboBox
        krankenkassenarten = ["------ Auswählen ------","gesetzlich", "freiwillig gesetzlich", "privat"]
        self.krankenk_var = tk.StringVar(self)
        self.combobox2 = Combobox(auswahl_Frame1, values=krankenkassenarten, textvariable=self.krankenk_var, width=24)
        self.combobox2.bind("<<ComboboxSelected>>", self.update_dentist_options)  # Bind the selection event
        self.combobox2.current(0)  # Set the default selection
        self.combobox2.place(rely=0.05, relx=0.56)
        # Dentist Selection Label
        self.dentist_label = tk.Label(self, text="Behandelter Zahnarzt:", background="white", font=("Arial", 10, "bold"))
        self.dentist_label.place(rely=0.1, relx=0.478)

        # Dentist ComboBox
        self.dentist_var = tk.StringVar(self)
        self.dentist_combobox = Combobox(self, textvariable=self.dentist_var, state="readonly", width=24)
        self.dentist_combobox.place(rely=0.1, relx=0.641)
        self.dentist_combobox.bind("<<ComboboxSelected>>", self.updateDescription)

        self.zahnarzt_button = tk.Button(self, text="Verfügbarkeit anzeigen",  width=22, bg="#3B6064", command=self.show_appointments, fg="white", font=("Arial", 9, "bold"))
        # self.zahnarzt_button.place(rely=0.145, relx=0.6415)
        self.zahnarzt_button.place(rely=0.430, relx=0.8270)
        # self.zahnarzt_button.pack_forget()

        self.cost_label = tk.Label(auswahl_Frame1, text="Behandlungskosten: $0.00", pady=20, background="#B2CFD2", height=3, padx=60, justify="center", font=("Arial", 10, "bold"))
        self.cost_label.place(rely=0.446, relx=0.4)

        self.buttonBerechnen = tk.Button(auswahl_Frame1, text="Kosten Berechnen",background="#3B6064", fg="white", command=self.calculate_cost, font=("Arial", 10, "bold"), padx=5, pady=10)
        self.buttonBerechnen.place(rely=0.8, relx=0.399)
        self.buttonBerechnen = tk.Button(auswahl_Frame1, text="Termin Buchen",background="#3B6064", fg="white", font=("Arial", 10, "bold"), padx=10, pady=10, command=self.termin_bestätigung)
        self.buttonBerechnen.place(rely=0.8, relx=0.5516)

        self.terminDescr = tk.Frame(auswahl_Frame1, background="white", width=266, height=280)
        self.terminDescr.place(rely=0.03, relx=0.74)

        self.description = tk.Label(self.terminDescr, text="keine Beschreibung", width=36, height=10, background="white")
        self.description.place(rely=0.24, relx=0.01)

        #--------Auswahl Frame unter---------
        auswahl_Frame3 = tk.LabelFrame(self, text="", height=237, background="white")
        auswahl_Frame3.place(rely=0.5395, relx=0.186)

        columns = ("Datum", "Uhrzeit", "Füllmaterial", "Zahnanzahl", "Arzt")
        self.treeview_gebu_Termin = ttk.Treeview(auswahl_Frame3, columns=columns, height=10, show='headings')
        self.treeview_gebu_Termin.pack()

        # Überschriften für den Treeview festlegen
        for column in columns:
            self.treeview_gebu_Termin.heading(column, text=column, anchor='center')
            self.treeview_gebu_Termin.column(column, width=210, anchor='center')

        self.update_treeview()
    
    def viewAktualisieren(self):
        self.update_treeview()

    
    def update_treeview(self):

        # Votre logique pour charger les données
        if self.krankenkasse_art == "Privat":
            excel_file_path = "Database\Patienten_Zahnärzte_Kosten.xlsx"
            selected_columns = ["Datum", "Uhrzeit", "Füllmaterial", "Anzahl Zähne", "Arzt"]
            df = pd.read_excel(excel_file_path, sheet_name="privat", usecols=selected_columns)

            if df is not None:
                # Insérer les données dans le ttk.Treeview
                for index, row in df.iterrows():
                    self.treeview_gebu_Termin.insert("", "end", values=tuple(row))
        elif self.krankenkasse_art == "gesetzlich":
            df = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="gesetzlich")
            if df is not None:
                # Insérer les données dans le ttk.Treeview
                for index, row in df.iterrows():
                    self.treeview_gebu_Termin.insert("", "end", values=tuple(row))
        elif self.krankenkasse_art == "freiwillig gesetzlich":
            df = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="freiwillig gesetzlich")
            if df is not None:
                # Insérer les données dans le ttk.Treeview
                for index, row in df.iterrows():
                    self.treeview_gebu_Termin.insert("", "end", values=tuple(row))

        style = ttk.Style()
        style.configure("treeview_gebu_Termin", rowheight=40)  # Augmenter la valeur de rowheight pour augmenter l'espace
        # self.update_image(self)  # Bildwechsel starten  
        self.update_time()  # Time Update

        # Variable pour stocker le rendez-vous sélectionné
        self.selection_rendezvous = None

    def open_image(self):
        file_path = filedialog.askopenfilename()  # Ouvre la boîte de dialogue pour sélectionner un fichier
        if file_path:
            image = Image.open(file_path)
            photo = ImageTk.PhotoImage(image)
            self.label_img.config(image=photo)
            self.label_img.image = photo  # Garde une référence à l'image pour éviter sa suppression

    def show_appointments(self):
        selected_dentist = self.dentist_var.get()
        if selected_dentist:
            appointments_window = tk.Toplevel(self.root)
            appointments_window.title(f"Terminkalender für {selected_dentist}")

            treeview = ttk.Treeview(appointments_window, columns=('Datum', 'Uhrzeit'))
            treeview.grid(row=0, column=0, padx=10, pady=10)

            for heading in ['Datum', 'Uhrzeit']:
                treeview.heading(heading, text=heading)

            appointments = self.zahnaerzte[selected_dentist]
            for appointment in appointments:
                date, time = self.parse_appointment(appointment)

                if 8 <= int(time.split(':')[0]) < 12 or 14 <= int(time.split(':')[0]) < 16:
                    treeview.insert('', 'end', values=(date, time))

            treeview.bind("<Double-1>", lambda event: self.select_appointment(treeview))

    def parse_appointment(self, appointment):
        self.date, self.time = appointment.split()
        return self.date, self.time

    def select_appointment(self, treeview):
        selection = treeview.selection()

        if selection:
            item = treeview.item(selection)
            self.date, self.time = item['values']

            # Stocker le rendez-vous dans la feuille 'privat' du fichier Excel
            self.save_appointment_to_excel(self.date, self.time)

    def save_appointment_to_excel(self, date, time):
        # t = teeth, f = filling, sk = selected_krankenk, p = problematik,  k = Konste
        krankenkasse = self.krankenk_var.get()
        t = self.teeth_var.get()
        f = self.fill_material_var.get()
        sk = self.krankenk_var.get()
        p = self.problematik_var.get()
        excel_file_path = 'Database\Patienten_Zahnärzte_Kosten.xlsx'
        name = self.patient_names
        self.dentist = self.dentist_var.get()
        try:
            # Charger le classeur Excel existant
            wb = load_workbook(excel_file_path)

            # Sélectionner la feuille 'privat' (créer si elle n'existe pas)
            if krankenkasse == "privat":
                if 'privat' not in wb.sheetnames:
                    wb.create_sheet('privat')
                ws = wb['privat']
                # Füge eine neue Zeile mit den Termindetails hinzu
                for row in ws.iter_rows(min_row=2, max_col=8, max_row=ws.max_row):
                    if row[5].value == date and row[6].value == time:
                        messagebox.showinfo("Termin existiert", f"Der Termin um {time} am {date} existiert bereits.")
                        return
                    else:
                        # ws.append([name, t, f, sk, p, date, time, self.dentist_var.get()])
                        ws.append([date, time, f, t, self.dentist_var.get(), name, p])

                # speichere die Änderungen
                wb.save(excel_file_path)
                print(f"Termin wurde in {excel_file_path} gespeichert, Sheet 'Privat'")
            elif krankenkasse == "gesetzlich":
                if 'gesetzlich' not in wb.sheetnames:
                    wb.create_sheet('gesetzlich')
                ws = wb['gesetzlich']
                 # Füge eine neue Zeile mit den Termindetails hinzu
                for row in ws.iter_rows(min_row=2, max_col=8, max_row=ws.max_row):
                    if row[5].value == date and row[6].value == time: # and row[7] == self.dentist
                        messagebox.showinfo("Termin existiert", f"Der Termin um {time} am {date} existiert bereits.")
                        return
                    else: 
                        ws.append([date, time, f, t, self.dentist_var.get(), name, p])

                # speichere die Änderungen
                wb.save(excel_file_path)
                print(f"Termin wurde in {excel_file_path} gespeichert, Sheet 'gesetzlich'")
            else:
                if 'freiwillig gesetzlich' not in wb.sheetnames:
                    wb.create_sheet('freiwillig gesetzlich')
                ws = wb['freiwillig gesetzlich']
                # Füge eine neue Zeile mit den Termindetails hinzu
                for row in ws.iter_rows(min_row=2, max_col=8, max_row=ws.max_row):
                    if row[5].value == date and row[6].value == time:
                        messagebox.showinfo("Termin existiert", f"Der Termin um {time} am {date} existiert bereits.")
                        return
                    else:
                        ws.append([date, time, f, t, self.dentist_var.get(), name, p])

                # Änderungen speichern
                wb.save(excel_file_path)
                print(f"Termin wurde in {excel_file_path} gespeichert, Sheet 'freiwillig gesetzlich'")
                
        except FileNotFoundError:
            print("File wurde nicht gefunden.")

    def open_change_password_window(self):
        change_password_window = ChangePassworWindow(self.root, self.current_user)
        change_password_window.title("Passwort ändern")

    def open_Kalender_view(self):
        self.kalender_view = TerminView(self)
        self.kalender_view.mainloop()

    def calculate_cost(self):
        teeth = int(self.teeth_var.get())
        filling = self.fill_material_var.get()
        selected_krankenk = self.krankenk_var.get()
        problematik = self.problematik_var.get()
        last_names2 = username

        # ## TODO: wenn Teeth == 0 or wenn filling == 0 or ....... messagebox.showerror("Fehler", "Anzahl teeth eingeben")
        # Correspondance des options de combobox avec les pourcentages
        if problematik == "Karies klein":
            self.anzeige_Zeit_Label.config(text="0.25")
            if filling == "normal":
                if selected_krankenk == "gesetzlich":
                    cost = 100 * teeth * 0.8
            elif filling == "höherwertig":
                    cost = 180 * teeth * 0.7
            elif filling == "höchstwertig":
                    cost = 250 * teeth * 0.5
        if problematik == "Karies klein":
            self.anzeige_Zeit_Label.config(text="0.25")
            if filling == "normal":
                if selected_krankenk == "privat":
                    cost = 100 * teeth * 1
            elif filling == "höherwertig":
                    cost = 180 * teeth * 1
            elif filling == "höchstwertig":
                    cost = 250 * teeth * 0.95
        if problematik == "Karies groß":
            self.anzeige_Zeit_Label.config(text="1")
            if filling == "normal":
                if selected_krankenk == "gesetzlich":
                    cost = 200 * teeth * 0.8
            elif filling == "höherwertig":
                if selected_krankenk == "gesetzlich":
                    cost = 360 * teeth * 0.7
            elif filling == "höchstwertig":
                    cost = 480 * teeth * 0.5
        if problematik == "Karies groß":
            if filling == "normal":
                self.anzeige_Zeit_Label.config(text="1")
                if selected_krankenk == "privat":
                    cost = 200 * teeth * 1
            elif filling == "höherwertig":
                    cost = 360 * teeth * 1
            elif filling == "höchstwertig":
                    cost = 480 * teeth * 0.95
        if problematik == "Teilkrone":
            if filling == "normal":
                self.anzeige_Zeit_Label.config(text="1")
                if selected_krankenk == "gesetzlich":
                    cost = 1000 * teeth * 0.8
            elif filling == "höherwertig":
                    cost = 1800 * teeth * 0.7
            elif filling == "höchstwertig":
                    cost = 3100 * teeth * 0.30
        if problematik == "Teilkrone":
            if filling == "normal":
                self.anzeige_Zeit_Label.config(text="1")
                if selected_krankenk == "privat":
                    cost = 1000 * teeth * 1
            elif filling == "höherwertig":
                    cost = 1800 * teeth * 0.95
            elif filling == "höchstwertig":
                    cost = 3100 * teeth * 0.85
        if problematik == "krone":
            if filling == "normal":
                self.anzeige_Zeit_Label.config(text="2")
                if selected_krankenk == "gesetzlich":
                    cost = 2800 * teeth * 0.7
            elif filling == "höherwertig":
                    cost = 3600 * teeth * 0.5
            elif filling == "höchstwertig":
                    cost = 4200 * teeth * 0.25
        if problematik == "krone":
            if filling == "normal":
                self.anzeige_Zeit_Label.config(text="2")
                if selected_krankenk == "privat":
                    cost = 2800 * teeth * 1
            elif filling == "höherwertig":
                    cost = 3600 * teeth * 0.85
            elif filling == "höchstwertig":
                    cost = 4200 * teeth * 0.75
        if problematik == "Wurzelbehandlung":
            self.anzeige_Zeit_Label.config(text="1.5")
            if filling == "normal":
                if selected_krankenk == "gesetzlich":
                    cost = 750 * teeth * 0.9
            elif filling == "höherwertig":
                    self.anzeige_Zeit_Label.config(text="1.5")
                    cost = 1200 * teeth * 0.70
            elif filling == "höchstwertig":
                    cost = 1500 * teeth * 0.45
                    self.anzeige_Zeit_Label.config(text="2")
        if problematik == "Wurzelbehandlung":
            if filling == "normal":
                if selected_krankenk == "privat":
                    cost = 750 * teeth * 1
            elif filling == "höherwertig":
                    cost = 1200 * teeth * 0.95
            elif filling == "höchstwertig":
                    cost = 1500 * teeth * 0.85
        
        self.cost_var.set("Behandlungskosten: {}€".format(cost))
        self.cost_label.config(textvariable=self.cost_var)

    def termin_bestätigung(self):

        teeth = int(self.teeth_var.get())
        filling = self.fill_material_var.get()
        selected_krankenk = self.krankenk_var.get()
        problematik = self.problematik_var.get()
        arzt = self.dentist_var.get()
        last_names2 = username

        # Arbeitsblatt mit openpyxl öffnen
        workbook = openpyxl.load_workbook('Database\Patienten_Zahnärzte_Kosten.xlsx')

        if selected_krankenk == "privat":
            sheet = workbook['privat']
            sheet['A1'] = 'Datum'
            sheet['B1'] = 'Uhrzeit'
            sheet['C1'] = 'Filling Material'
            sheet['D1'] = 'Anzahl zu behandelnder Zähne'
            sheet['E1'] = 'Arzt'
            sheet['F1'] = 'Patient'
            sheet['G1'] = 'Problematik'		

            # Werte
            # Überprüfung, ob der Arzt bereits einen Termin an diesem Datum hat
            found_duplicate = False
            for row in range(2, sheet.max_row + 1):  # Starte bei Zeile 2 (Annahme von Überschriften in Zeile 1)
                if sheet[f'A{row}'].value == self.date and sheet[f'E{row}'].value == arzt:
                    found_duplicate = True
                    break
            # Wenn ein Termin für diesen Arzt an diesem Datum gefunden wurde, zeige eine entsprechende Nachricht
            if found_duplicate:
                messagebox.showerror(f'Der Arzt {arzt} hat bereits einen Termin am {self.date}.')
            else:
                row = sheet.max_row + 1  # Nächste verfügbare Zeile
                sheet[f'A{row}'] = self.date
                sheet[f'B{row}'] = self.time
                sheet[f'C{row}'] = filling
                sheet[f'D{row}'] = teeth
                sheet[f'E{row}'] = arzt
                sheet[f'F{row}'] = last_names2
                sheet[f'G{row}'] = problematik
                # Datei speichern
                workbook.save('Database\Patienten_Zahnärzte_Kosten.xlsx')
        elif selected_krankenk == "gesetzlich":
            sheet = workbook['gesetzlich']
            sheet['A1'] = 'Datum'
            sheet['B1'] = 'Uhrzeit'
            sheet['C1'] = 'Filling Material'
            sheet['D1'] = 'Anzahl zu behandelnder Zähne'
            sheet['E1'] = 'Arzt'
            sheet['F1'] = 'Patient'
            sheet['G1'] = 'Problematik'																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																								
            # Werte
            # Überprüfung, ob der Arzt bereits einen Termin an diesem Datum hat
            found_duplicate = False
            for row in range(2, sheet.max_row + 1):  # Starte bei Zeile 2 (Annahme von Überschriften in Zeile 1)
                if sheet[f'A{row}'].value == self.date and sheet[f'E{row}'].value == arzt:
                    found_duplicate = True
                    # print("Error Beim Termin")
                    break
            # Wenn ein Termin für diesen Arzt an diesem Datum gefunden wurde, zeige eine entsprechende Nachricht
            if found_duplicate:
                messagebox.showerror(f'Der Arzt {arzt} hat bereits einen Termin am {self.date}.')
            else:
                row = sheet.max_row + 1  # Nächste verfügbare Zeile
                sheet[f'A{row}'] = self.date
                sheet[f'B{row}'] = self.time
                sheet[f'C{row}'] = filling
                sheet[f'D{row}'] = teeth
                sheet[f'E{row}'] = arzt
                sheet[f'F{row}'] = last_names2
                sheet[f'G{row}'] = problematik
                # Datei speichern
                workbook.save('Database\Patienten_Zahnärzte_Kosten.xlsx')
        else:
            sheet = workbook['freiwillig gesetzlich']
            sheet['A1'] = 'Datum'
            sheet['B1'] = 'Uhrzeit'
            sheet['C1'] = 'Filling Material'
            sheet['D1'] = 'Anzahl zu behandelnder Zähne'
            sheet['E1'] = 'Arzt'
            sheet['F1'] = 'Patient'
            sheet['G1'] = 'Problematik'																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																																								
            # Werte
            # Überprüfung, ob der Arzt bereits einen Termin an diesem Datum hat
            found_duplicate = False
            for row in range(2, sheet.max_row + 1):  # Starte bei Zeile 2 (Annahme von Überschriften in Zeile 1)
                if sheet[f'A{row}'].value == self.date and sheet[f'E{row}'].value == arzt:
                    found_duplicate = True
                    break
            # Wenn ein Termin für diesen Arzt an diesem Datum gefunden wurde, zeige eine entsprechende Nachricht
            if found_duplicate:
                messagebox.showerror(f"Der Arzt {arzt} hat bereits einen Termin am {self.date}.")
            else:
                row = sheet.max_row + 1  # Nächste verfügbare Zeile
                sheet[f'A{row}'] = self.date
                sheet[f'B{row}'] = self.time
                sheet[f'C{row}'] = filling
                sheet[f'D{row}'] = teeth
                sheet[f'E{row}'] = arzt
                sheet[f'F{row}'] = last_names2
                sheet[f'G{row}'] = problematik
                # Datei speichern
                workbook.save('Database\Patienten_Zahnärzte_Kosten.xlsx')
             
    def update_dentist_options(self, event):
        selected_krankenk = self.krankenk_var.get()
        if selected_krankenk == "gesetzlich":
            allowed_dentists = ["Herr Dr. Kraft", "Herr Dr. Hausmann"]
        elif selected_krankenk == "privat":
            allowed_dentists = ["Herr Dr. Huber", "Frau Dr. Wurzel", "Frau Dr. Winkel"]
        elif selected_krankenk == "freiwillig gesetzlich":
            allowed_dentists = ["Frau Dr. Winkel", "Herr Dr. Hausmann"]
        else:
            allowed_dentists = ["Krankenkasse Auswählen"]

        self.dentist_var.set("")  # Clear the current selection
        self.dentist_combobox['values'] = allowed_dentists  # Update the ComboBox widget


    def updateDescription(self, event):
         selected_Artz = self.dentist_combobox.get()
         if selected_Artz == "Herr Dr. Huber":
            self.description.config(background="#F8C9CF", text="Öffnungszeiten von Dr. Huber:\nMo-Fr: 8-12 Uhr und 14-16 Uhr\n\n Drücken sie bitte ganz unter\n auf Terminkalender")
         if selected_Artz == "Herr Dr. Kraft":
              self.description.config(background="#F8C9CF", text="Öffnungszeiten von Dr. Kraft:\nMo-Fr: 10-12 Uhr und 14-18 Uhr\n\n Drücken sie bitte ganz unter\n auf Terminkalender")
         if selected_Artz == "Frau Dr. Winkel":
              self.description.config(background="#F8C9CF", text="Öffnungszeiten von Dr. Winkel:\nMo-Fr: 8-14 Uhr\n\n Drücken sie bitte ganz unter\n auf Terminkalender")
         if selected_Artz == "Herr Dr. Hausmann":
              self.description.config(background="#F8C9CF", text="Öffnungszeiten von Dr. Hausmann:\nMo: 9-12 Uhr und Fr 12-16 Uhr\n\n Drücken sie bitte ganz unter\n auf Terminkalender")
         if selected_Artz == "Frau Dr. Wurzel":
              self.description.config(background="#F8C9CF", text="Öffnungszeiten von Dr. Wurzel:\nDi: 8-12 und Do: 12-18 Uhr\n\n Drücken sie bitte ganz unter\n auf Terminkalender")

    def update_time(self):
        current_datetime = datetime.now()
        current_time = datetime.now().strftime("%H:%M:%S")  # Aktuelle Uhrzeit abrufen
        current_date = current_datetime.strftime("%d.%m.%Y")  # Aktuelles Datum abrufen
        self.current_time_label.config(text=current_time)  # Uhrzeit im Label aktualisieren

        datetime_text = f"{current_date} {current_time}"  # Kombinierten Text erstellen
        self.current_time_label.config(text=datetime_text)  # Text im Label aktualisieren

        self.current_time_label.after(1000, self.update_time)  # Nach 1000 Millisekunden erneut aufrufen

    def logout(self):
        self.destroy()
        self.master.deiconify()
        self.master.username_entry.delete(0, tk.END)
        self.master.password_entry.delete(0, tk.END)

    def animate_label(self):
        current_fg = self.firma_name_label["foreground"]
        if current_fg == "white":
            self.firma_name_label["foreground"] = "black"
        else:
            self.firma_name_label["foreground"] = "white"
        self.firma_name_label.after(500, self.animate_label)  # Nach 500 Millisekunden erneut aufrufen

class TerminView(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Terminkalender")
        self.geometry("525x295")
        self.config(bg="#3B6064")
        self.maxsize(525, 355)
        self.minsize(525, 355)

        #variable aus andere classe
        self.fill_material = self.patient_view.fill_material_var
        dentist = self.patient_view.dentist_var
        self.problematik = self.patient_view.problematik_var

        # self.teeth_var = tk.StringVar()
        # self.cost_var = tk.StringVar()

        self.selected_datetime = tk.StringVar()
        self.selected_datetime_label = tk.Label(self, textvariable=self.selected_datetime)
        self.selected_datetime_label.pack()

        self.date_frame = tk.Frame(self)
        self.date_frame.pack()

        self.calendar = Calendar(self.date_frame, selectmode="day", year=2023, month=10, day=25)
        self.calendar.pack(side="left")

        self.time_frame = tk.Frame(self)
        self.time_frame.pack()

        self.hour_spinbox = ttk.Spinbox(self.time_frame, from_=0, to=23, width=2)
        self.hour_spinbox.set(12)  # Heure par défaut
        self.hour_spinbox.pack(side="left")
        self.hour_label = tk.Label(self.time_frame, text=":")
        self.hour_label.pack(side="left")
        self.minute_spinbox = ttk.Spinbox(self.time_frame, from_=0, to=59, width=2)
        self.minute_spinbox.set(0)  # Minutes par défaut
        self.minute_spinbox.pack(side="left")

        self.date_frame.bind("<<CalendarSelected>>", self.on_date_select)
        self.hour_spinbox.bind("<FocusOut>", self.on_time_select)
        self.minute_spinbox.bind("<FocusOut>", self.on_time_select)


        self.buch_button = tk.Button(self, text="Termin bestätigen", background="white", fg="black", font=("Arial", 13), command=self.select_button_click)
        self.buch_button.pack(side="bottom",  padx=5, pady=10, ipadx=50)

        self.update_selected_datetime()

    def update_selected_datetime(self):
        selected_date = self.calendar.get_date()
        selected_time = f"{self.hour_spinbox.get()}:{self.minute_spinbox.get()}"
        self.selected_datetime.set(f"Date et Heure sélectionnées: {selected_date} {selected_time}")

    def on_date_select(self, event):
        self.update_selected_datetime()

    def on_time_select(self, event):
        self.update_selected_datetime()

    def select_button_click(self):
        selected_date = self.calendar.get_date()
        selected_time = f"{self.hour_spinbox.get()}:{self.minute_spinbox.get()}"

        df = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx")
        self.material = self.fill_material
        print(f"Date sélectionnée: {selected_date}")
        print(f"Heure sélectionnée: {selected_time}")
        print(f"Heure sélectionnée: {self.material}")

class DentistView(tk.Toplevel):
    def __init__(self, master, last_names):
        super().__init__(master)
        self.geometry("1300x580")
        self.minsize(1300, 580)
        self.maxsize(1300, 580)
        self.title("Zahnarztansicht")
        self.config(background="#3B6064")
        self.last_names2 = last_names

        # Zielgröße für das redimensionierte Bild
        self.width = 530
        self.height = 260

        self.untersten_Frame = tk.Frame(self, width=1300, height=30, background="#5B949A")
        self.untersten_Frame.pack(side="bottom", fill="y")
        
        self.firma_name_label = tk.Label(self.untersten_Frame, text="Zahnärzte am Park GmbH", background="#5B949A", fg="white", font=("Arial", 11, "bold"))
        # firma_name_label.pack(side="left", fill="x")
        self.firma_name_label.place(x=15, y=5)
        self.animate_label()  # Animation starten

        self.current_time_label = tk.Label(self.untersten_Frame, text="", font=("Arial", 11, "bold"), fg="white", bg="#5B949A")
        self.current_time_label.place(x=1150, y=5)

        self.NameLabel = tk.Label(self.untersten_Frame, text= username, font=("Arial", 11, "bold"), background="#5B949A", fg="white")
        self.NameLabel.place(x=600, y=5)

        self.termin_Frame = tk.LabelFrame(self, text="Terminkalender",border=0, background="white")
        self.termin_Frame.place(x=7, y=7)

        columns=("Patient", "Problematik", "Filling Material", "Anzahl Zähne", "Datum", "Uhrzeit")
        self.treeview_gebu_Termin = ttk.Treeview(self.termin_Frame, columns= columns, height=11, show='headings')
        self.treeview_gebu_Termin.pack()

        # Überschrift für den Treeview festlegen
        self.treeview_gebu_Termin.heading("Patient", text="Patientname", anchor='center')
        self.treeview_gebu_Termin.heading("Problematik", text="Behandlungen", anchor='center')
        self.treeview_gebu_Termin.heading("Filling Material", text="Material", anchor='center')
        self.treeview_gebu_Termin.heading("Anzahl Zähne", text="Zähne", anchor='center')
        self.treeview_gebu_Termin.heading("Datum", text="Datum", anchor='center')
        self.treeview_gebu_Termin.heading("Uhrzeit", text="Uhrzeit", anchor='center')

        self.treeview_gebu_Termin.column("Patient", width=150)
        self.treeview_gebu_Termin.column("Problematik", width=150)
        self.treeview_gebu_Termin.column("Filling Material", width=150)
        self.treeview_gebu_Termin.column("Anzahl Zähne", width=80)
        self.treeview_gebu_Termin.column("Datum", width=110)
        self.treeview_gebu_Termin.column("Uhrzeit", width=110)

        # Überschriften für den Treeview festlegen
        for column in columns:
            self.treeview_gebu_Termin.heading(column, text=column, anchor='center')
        # Spalten konfigurieren, um die Werte zu zentrieren
        for column in columns:
            self.treeview_gebu_Termin.column(column, anchor='center')
        style = ttk.Style()
        style.configure('treeview_gebu_Termin', rowheight=30)  # Hier können Sie die gewünschte Höhe anpassen

        self.imageView_frame = tk.Frame(self, height=260, width=530)
        self.imageView_frame.place(x=765, y=7)

        df = None  # Initialisation par défaut de df
        if username == "Huber" or username == "Wurzel":
            df = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="privat")
        elif username == "Kraft":
             df = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="gesetzlich")
        # elif username == "Winkel":
        elif username == "Hausmann":
             df1 = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="gesetzlich")
             df2 = pd.read_excel("Database\Patienten_Zahnärzte_Kosten.xlsx", sheet_name="freiwillig gesetzlich")
             df = pd.concat([df1, df2])
             
        if df is not None:
            font = Font(size=11, weight="bold")
            for index, row in df.iterrows():
                self.treeview_gebu_Termin.insert("", "end", values=(row[5], row[6], row[2], row[3], row[0], row[1]), tags="custom_font")
            self.treeview_gebu_Termin.tag_configure("custom_font", font=font)
        style = ttk.Style()
        style.configure("treeview_gebu_Termin", rowheight=40)  # Augmenter la valeur de rowheight pour augmenter l'espace


        # Zielgröße für das redimensionierte Bild
        self.width = 530
        self.height = 260
        # Liste der Bildpfade
        self.image_paths = ["Bilder/img1.jpg", "Bilder/img2.jpg", "Bilder/img3.jpg", "Bilder/img4.jpg", "Bilder/img5.jpg"]

        self.current_image_index = 0  # Index des aktuellen Bildes

        # Hintergrundbild laden und redimensionieren
        image = Image.open(self.image_paths[self.current_image_index])
        image = image.resize((self.width, self.height), Image.LANCZOS)

        # Redimensioniertes Hintergrundbild in ein Tkinter-Bildobjekt konvertieren
        self.background_image = ImageTk.PhotoImage(image)

        # Hintergrundbild-Label erstellen und platzieren
        self.background_label = tk.Label(self.imageView_frame, image=self.background_image)
        self.background_label.pack()

        # OPtionen (Button)
        self.optionFrame1 = tk.Frame(self, width=351, height=256, bg="#3B6064", highlightbackground="#5B949A", highlightthickness=0.5)
        self.optionFrame1.place(x=765, y=278)
        #----------------------------
        # self.logout_button = tk.Button(self.optionFrame1, text="Krankenkasse\nändern", width=18, height=2, font=("Arial", 9, "bold"))
        # self.logout_button.place(x=18, y=25)
        self.image_frame = tk.Frame(self.optionFrame1, width=317, height=162)
        self.image_frame.place(rely=0.06, relx=0.0480)
        # self.logout_button = tk.Button(self.optionFrame1, text=" ", width=44, height=10, font=("Arial", 9, "bold"))
        # self.logout_button.place(x=18, y=15)
        self.logout_button = tk.Button(self.optionFrame1, text="Behandlungszeit\nändern", width=20, height=2, font=("Arial", 9, "bold"), background='#5B949A', foreground='white', command=self.set_behandlungzeit)
        self.logout_button.place(x=18, y=200)
        self.logout_button = tk.Button(self.optionFrame1, text="Krankenkasse\nändern", width=20, height=2, font=("Arial", 9, "bold"), background='#5B949A', foreground='white', command=self.set_krankenkasse)
        self.logout_button.place(x=187, y=200)

        # Info Frame
        self.Name_Label = tk.Label(self.image_frame ,text="Titel: Herr", font=("Arial", 11, "bold"))
        self.Name_Label.place(rely=0.08, relx=0.03)
        self.Name_Label = tk.Label(self.image_frame ,text="Name: "+ last_names, font=("Arial", 11, "bold"))
        self.Name_Label.place(rely=0.30, relx=0.03)
        self.Krankenkass_Label = tk.Label(self.image_frame ,text="Krankenkasse: xxxx", font=("Arial", 11, "bold"))
        self.Krankenkass_Label.place(rely=0.53, relx=0.03)

        # Charger l'image
        image = Image.open("sign-out-alt.png")  # Mettez ici le chemin vers votre image
        # image = Image.open("cross.png")  # Mettez ici le chemin vers votre image
        # Redimensionner l'image si nécessaire
        image = image.resize((2,2),  Image.LANCZOS)  
        icon = ImageTk.PhotoImage(image)

        #----------------------------
        #----------------------------
        self.optionFrame2 = tk.Frame(self, width=170, height=256, bg="#3B6064", highlightbackground="#5B949A", highlightthickness=0.5)
        self.optionFrame2.place(x=1126, y=278)
        #----------------------------
        self.logout_button = tk.Button(self.optionFrame2, text="Passwort ändern", width=18, height=2, font=("Arial", 9, "bold"), background='#5B949A', foreground='white', command=self.open_change_password_D_window)
        self.logout_button.place(x=18, y=16)
        #----------------------------
        self.Hintergrung_button = tk.Button(self.optionFrame2, text="Hintergrung", width=18, height=2, font=("Arial", 9, "bold"), background='#5B949A', foreground='white', command=self.set_hintergrund)
        self.Hintergrung_button.place(x=18, y=74.2)
        #----------------------------
        self.Sprache_button = tk.Button(self.optionFrame2, text="Englisch/Deutsch", width=18, height=2, font=("Arial", 9, "bold"), background='#5B949A', foreground='white', command=self.set_language)
        self.Sprache_button.place(x=18, y=135)
        #----------------------------
        self.logout_button = tk.Button(self.optionFrame2, text="Auslogen", command=self.logout, width=16, height=2, background="#E36370", fg="white", font=("Arial", 10, "bold"))
        self.logout_button.place(x=18, y=200)

        self.update_image()  # Bildwechsel starten
        self.update_time()  # Time Update

        self.optionFrame3 = tk.Frame(self, width=751, height=256, bg="#3B6064", highlightbackground="#5B949A", highlightthickness=0.5)
        self.optionFrame3.place(x=7.5, y=278)

        self.count_patient()
        self.warteZeit_Label = tk.Label(self.optionFrame3, text="Terminanzahl: ", width=22, height=5, fg="black", font=("Arial", 11, "bold"))
        self.warteZeit_Label.place(x=15, y=12)
        self.TerminAnzahl_Label = tk.Label(self.optionFrame3, text="Patientanzahl: " + str(self.patientanzahl), width=28, height=5, fg="black", font=("Arial", 11, "bold"))
        self.TerminAnzahl_Label.place(x=218, y=12)
        anzahlPatien_button = tk.Button(self.optionFrame3, text="Patientlist", width=22, height=3, fg="white", background='#5B949A', font=("Arial", 11, "bold"))
        anzahlPatien_button.place(x=530, y=12)
        Termin_Hinzuf_button = tk.Button(self.optionFrame3, text="Termin Hinzufügen", width=22, height=3, fg="white", background='#5B949A', font=("Arial", 11, "bold"))
        Termin_Hinzuf_button.place(x=530, y= 90)
        Termin_Hinzuf_button = tk.Button(self.optionFrame3, text="Patient Hinzufügen", width=22, height=3, fg="white", background='#5B949A', font=("Arial", 11, "bold"))
        Termin_Hinzuf_button.place(x=530, y= 170)

    # Funktion zum Aktualisieren des Bildes
    def count_patient(self):
        df = pd.read_excel('Database/Patienten_Zahnärzte_Kosten.xlsx', sheet_name='Stamm-Patienten2')
        self.patientanzahl = df['Patient'].count()
        return self.patientanzahl

    def update_image(self):
        # Nächsten Bildindex berechnen
        self.current_image_index = (self.current_image_index + 1) % len(self.image_paths)

        # Bild laden und anzeigen
        image = Image.open(self.image_paths[self.current_image_index])
        image = image.resize((self.width, self.height), Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        self.background_label.config(image=photo)
        self.background_label.image = photo  # Referenz behalten, um das Bild im Speicher zu halten

        # Timer für das nächste Bild setzen (nach 4 Sekunden)
        self.after(3000, self.update_image)

    def update_time(self):
        current_datetime = datetime.now()
        current_time = datetime.now().strftime("%H:%M:%S")  # Aktuelle Uhrzeit abrufen
        current_date = current_datetime.strftime("%d.%m.%Y")  # Aktuelles Datum abrufen
        self.current_time_label.config(text=current_time)  # Uhrzeit im Label aktualisieren

        datetime_text = f"{current_date} {current_time}"  # Kombinierten Text erstellen
        self.current_time_label.config(text=datetime_text)  # Text im Label aktualisieren

        self.current_time_label.after(1000, self.update_time)  # Nach 1000 Millisekunden erneut aufrufen

    def animate_label(self):
        current_fg = self.firma_name_label["foreground"]
        if current_fg == "white":
            self.firma_name_label["foreground"] = "black"
        else:
            self.firma_name_label["foreground"] = "white"
        self.firma_name_label.after(500, self.animate_label)  # Nach 500 Millisekunden erneut aufrufen

    def logout(self):
        self.destroy()
        self.master.deiconify()
        self.master.username_entry.delete(0, tk.END)
        self.master.password_entry.delete(0, tk.END)

    def open_change_password_D_window(self):
        change_password_window = ChangePasswordDenWindow(self.master, self.last_names2)
        change_password_window.title("Passwort ändern")

    def set_hintergrund(self):
        farbe = colorchooser.askcolor(title="Farbe auswählen")[1]
        self.config(bg=farbe)
    
    def set_language(self):
        print("language ist noch nicht fertig")
    
    def set_krankenkasse(self):
        print("krankenkasse ist noch nicht fertig")
    
    def set_behandlungzeit(self):
        print("zeit ist noch nicht fertig")

class ChangePassworWindow(tk.Toplevel):
    def __init__(self, parent, current_user):#
        super().__init__()#
        self.parent = parent #
        self.title("Passwort Änderung")
        self.geometry("500x400")
        self.config(background="#5B949A")
        self.resizable(False, False)
        self.current_user = current_user # Récupérer le nom d'utilisateur actuel

        self.old_password_label = tk.Label(self, text="Alte Passwort:", background="#5B949A", font=("Arial", 10, "bold"))
        self.old_password_label.place(rely=0.15, relx=0.18)
        self.old_password_entry = tk.Entry(self, show="*")
        self.old_password_entry.place(rely=0.15, relx=0.46)

        self.new_password_label = tk.Label(self, text="Neue Passwort:", background="#5B949A", font=("Arial", 10, "bold"))
        self.new_password_label.place(rely=0.25, relx=0.18)
        self.new_password_entry = tk.Entry(self, show="*")
        self.new_password_entry.place(rely=0.25, relx=0.46)

        self.confirm_password_label = tk.Label(self, text="Pass Bestätigen:", background="#5B949A", font=("Arial", 10, "bold"))
        self.confirm_password_label.place(rely=0.37, relx=0.18)
        self.confirm_password_entry = tk.Entry(self, show="*")
        self.confirm_password_entry.place(rely=0.37, relx=0.46)

        self.change_button = tk.Button(self, text="Bestätigen", background="white", command=self.change_password, padx= 40, pady=10)
        self.change_button.place(rely=0.55, relx=0.35)

    def change_password(self):
        old_password = self.old_password_entry.get()
        confirm_password = self.confirm_password_entry.get()
        new_password = self.new_password_entry.get()

        # Vérification si les champs de mot de passe sont vides
        if not old_password or not new_password or not confirm_password:
            messagebox.showerror("Fehler", "Bitte Felder ausfüllen!!")
            return

        # Vérification si le nouveau mot de passe correspond à la confirmation
        if new_password != confirm_password:
            messagebox.showerror("Fehler", "Das neue Passwort stimmt nicht mit der Bestätigung überein.")
            return

        # Lire le fichier Excel pour vérifier la connexion
        try:
            file_path = "Database/Patienten_Zahnärzte_Kosten.xlsx"
            workbook = openpyxl.load_workbook(file_path)
            sheet_name = "Stamm-Patienten2"
            sheet = workbook[sheet_name]

            # Créer une liste pour stocker les données mises à jour
            updated_data = []

            # Trouver la ligne correspondant au patient actuel
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if self.current_user.lower() in str(row[0]).lower():  # Vérifiez le nom du patient
                    stored_password = row[1]  # Récupérez le mot de passe stocké

                    if old_password == stored_password:
                        # Mettre à jour le mot de passe dans la liste
                        updated_data.append((row[0], new_password, *row[2:]))
                        messagebox.showinfo("Erfolg", "Password wurde aktualisiert")
                        self.destroy()
                        break
                    else:
                        messagebox.showerror("Fehler", "Altes Passwort stimmt nicht überein")
                        return
                else:
                    updated_data.append(row)

            # Écrire les données mises à jour dans la feuille Excel
            for row_idx, row_data in enumerate(updated_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            workbook.save(file_path)
        except Exception as e:
            print({str(e)})

class ChangePasswordDenWindow(tk.Toplevel):
    def __init__(self, parent, current_user):
        super().__init__()#
        self.parent = parent #
        self.title("Passwort Änderung")
        self.geometry("500x400")
        self.config(background="#5B949A")
        self.resizable(False, False)
        self.current_user = current_user # Récupérer le nom d'utilisateur actuel

        self.old_password_label = tk.Label(self, text="Alte Passwort:", background="#5B949A", font=("Arial", 10, "bold"))
        self.old_password_label.place(rely=0.15, relx=0.18)
        self.old_password_entry = tk.Entry(self, show="*")
        self.old_password_entry.place(rely=0.15, relx=0.46)

        self.new_password_label = tk.Label(self, text="Neue Passwort:", background="#5B949A", font=("Arial", 10, "bold"))
        self.new_password_label.place(rely=0.25, relx=0.18)
        self.new_password_entry = tk.Entry(self, show="*")
        self.new_password_entry.place(rely=0.25, relx=0.46)

        self.confirm_password_label = tk.Label(self, text="Pass Bestätigen:", background="#5B949A", font=("Arial", 10, "bold"))
        self.confirm_password_label.place(rely=0.37, relx=0.18)
        self.confirm_password_entry = tk.Entry(self, show="*")
        self.confirm_password_entry.place(rely=0.37, relx=0.46)

        self.change_button = tk.Button(self, text="Bestätigen", background="white", command=self.change_password, padx= 40, pady=10)
        self.change_button.place(rely=0.55, relx=0.35)

    def change_password(self):
        old_password = self.old_password_entry.get()
        confirm_password = self.confirm_password_entry.get()
        new_password = self.new_password_entry.get()

        # Vérification si les champs de mot de passe sont vides
        if not old_password or not new_password or not confirm_password:
            messagebox.showerror("Fehler", "Bitte Felder ausfüllen!!")
            return

        # Vérification si le nouveau mot de passe correspond à la confirmation
        if new_password != confirm_password:
            messagebox.showerror("Fehler", "Das neue Passwort stimmt nicht mit der Bestätigung überein.")
            return

        # Lire le fichier Excel pour vérifier la connexion
        try:
            file_path = "Database/Patienten_Zahnärzte_Kosten.xlsx"
            workbook = openpyxl.load_workbook(file_path)
            sheet_name = "Zahnärzte"  # Assurez-vous de la bonne feuille
            sheet = workbook[sheet_name]

            # Créer une liste pour stocker les données mises à jour
            updated_data = []

            # Trouver la ligne correspondant au patient actuel
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if self.current_user.lower() in str(row[0]).lower():  # Vérifiez le nom du patient
                    stored_password = row[1]  # Récupérez le mot de passe stocké

                    if old_password == stored_password:
                        # Mettre à jour le mot de passe dans la liste
                        updated_data.append((row[0], new_password, *row[2:]))
                        messagebox.showinfo("Erfolg", "Password wurde aktualisiert")
                        self.destroy()
                        break
                    else:
                        messagebox.showerror("Fehler", "Altes Passwort stimmt nicht überein")
                        
                        return
                else:
                    updated_data.append(row)

            # Écrire les données mises à jour dans la feuille Excel
            for row_idx, row_data in enumerate(updated_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            workbook.save(file_path)
        except Exception as e:
            # messagebox.showerror("Error", f"Une erreur s'est produite : {str(e)}")
            print({str(e)})


# Hauptprogramm
if __name__ == "__main__":
    loginWindow = LoginWindow()
    loginWindow.geometry("550x500")
    loginWindow.maxsize(550, 550)
    loginWindow.minsize(550, 550)

    loginWindow.mainloop()
