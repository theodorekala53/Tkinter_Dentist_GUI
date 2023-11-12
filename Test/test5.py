import tkinter as tk
from tkinter import messagebox
import openpyxl

class ChangePasswordApp:
    def __init__(self, patient_id):
        self.patient_id = patient_id

        self.root = tk.Tk()
        self.root.title("Changer le mot de passe")

        self.old_password_label = tk.Label(self.root, text="Ancien mot de passe:")
        self.old_password_entry = tk.Entry(self.root, show="*")
        
        self.new_password_label = tk.Label(self.root, text="Nouveau mot de passe:")
        self.new_password_entry = tk.Entry(self.root, show="*")

        self.confirm_new_password_label = tk.Label(self.root, text="Confirmer le nouveau mot de passe:")
        self.confirm_new_password_entry = tk.Entry(self.root, show="*")

        self.change_password_button = tk.Button(self.root, text="Changer le mot de passe", command=self.change_password)

        self.old_password_label.pack()
        self.old_password_entry.pack()
        self.new_password_label.pack()
        self.new_password_entry.pack()
        self.confirm_new_password_label.pack()
        self.confirm_new_password_entry.pack()
        self.change_password_button.pack()

    def change_password(self):
        old_password = self.old_password_entry.get()
        new_password = self.new_password_entry.get()
        confirm_new_password = self.confirm_new_password_entry.get()

        if new_password == confirm_new_password:
            if update_password(self.patient_id, old_password, new_password):
                messagebox.showinfo("Changement de mot de passe", "Mot de passe changé avec succès!")
            else:
                messagebox.showerror("Erreur", "Ancien mot de passe incorrect ou patient introuvable.")
        else:
            messagebox.showerror("Erreur", "Les nouveaux mots de passe ne correspondent pas.")

def update_password(patient_id, old_password, new_password):
    excel_file = "Patienten_Zahnärzte_Kosten.xlsx"  # Remplacez cela par le chemin réel de votre fichier Excel.
    wb = openpyxl.load_workbook(excel_file)
    sheet_name = "Stamm-Patienten2"
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row):
        if row[0].value == patient_id:
            stored_password = row[1].value

            if stored_password == old_password:
                row[1].value = new_password
                wb.save(excel_file)
                return True

    return False

if __name__ == "__main__":
    patient_id = "Meyer"  # Remplacez cela par la valeur réelle du patient connecté
    app = ChangePasswordApp(patient_id)
    app.root.mainloop()
