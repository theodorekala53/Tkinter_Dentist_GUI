# import tkinter as tk
# from tkinter import ttk

# class ZahnarztTerminGUI:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Zahnarzt Terminbuchung")

#         # Simulierte Daten - Zahnärzte und ihre Termine
#         self.zahnaerzte = {
#             "Dr. Müller": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
#             "Dr. Schmidt": ["2023-11-10 11:00", "2023-11-10 15:30", "2023-11-10 16:00"]
#         }

#         self.auswahl_frame = ttk.Frame(self.root)
#         self.auswahl_frame.pack(padx=20, pady=20)

#         self.zahnarzt_label = ttk.Label(self.auswahl_frame, text="Wählen Sie einen Zahnarzt:")
#         self.zahnarzt_label.grid(row=0, column=0, padx=10, pady=10)

#         self.zahnarzt_var = tk.StringVar()
#         self.zahnarzt_combobox = ttk.Combobox(self.auswahl_frame, textvariable=self.zahnarzt_var, values=list(self.zahnaerzte.keys()))
#         self.zahnarzt_combobox.grid(row=0, column=1, padx=10, pady=10)

#         self.zahnarzt_button = ttk.Button(self.auswahl_frame, text="Zahnarzt auswählen", command=self.anzeigen_terminkalender)
#         self.zahnarzt_button.grid(row=0, column=2, padx=10, pady=10)

#         # Hält eine Referenz zum aktuellen Terminkalender-Frame
#         self.terminkalender_frame = None

#     def anzeigen_terminkalender(self):
#         ausgewaehlter_zahnarzt = self.zahnarzt_var.get()
#         if ausgewaehlter_zahnarzt:
#             # Wenn es bereits einen Terminkalender gibt, entfernen Sie ihn
#             if self.terminkalender_frame:
#                 self.terminkalender_frame.destroy()

#             # Erstellen Sie einen neuen Terminkalender
#             self.terminkalender_frame = ttk.Frame(self.root)
#             self.terminkalender_frame.pack(padx=20, pady=20)

#             terminkalender_label = ttk.Label(self.terminkalender_frame, text=f"Terminkalender für {ausgewaehlter_zahnarzt}:")
#             terminkalender_label.grid(row=0, column=0, padx=10, pady=10)

#             terminkalender_treeview = ttk.Treeview(self.terminkalender_frame, columns=('Datum', 'Uhrzeit'))
#             terminkalender_treeview.grid(row=1, column=0, padx=10, pady=10)

#             terminkalender_treeview.heading('Datum', text='Datum')
#             terminkalender_treeview.heading('Uhrzeit', text='Uhrzeit')

#             # Füllen Sie den Terminkalender mit Daten des ausgewählten Zahnarztes
#             termine = self.zahnaerzte[ausgewaehlter_zahnarzt]
#             for termin in termine:
#                 terminkalender_treeview.insert('', 'end', values=(termin.split()[0], termin.split()[1]))

# if __name__ == "__main__":
#     root = tk.Tk()
#     app = ZahnarztTerminGUI(root)
#     root.mainloop()



        # # Simulierte Daten - Zahnärzte und ihre Termine
        # self.zahnaerzte = {
        #     "Herr Dr. Kraft": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
        #     "Herr Dr. Hausmann": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
        #     "Frau Dr. Winkel": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
        #     "Herr Dr. Huber": ["2023-11-10 09:00", "2023-11-10 10:30", "2023-11-10 14:00"],
        #     "Frau Dr. Wurzel": ["2023-11-10 11:00", "2023-11-10 15:30", "2023-11-10 16:00"]
        # }



# import tkinter as tk
# from tkinter import ttk
# import datetime

# class ZahnarztTerminGUI:
#     def __init__(self, root):
#         self.root = root
#         self.root.title("Zahnarzt Terminbuchung")

#         # Simuler des données - Dentistes et leurs plages de rendez-vous pour une année entière
#         self.zahnaerzte = {
#             "Dr. Müller": self.generer_rendezvous_pour_annee(),
#             "Dr. Schmidt": self.generer_rendezvous_pour_annee()
#         }

#         self.auswahl_frame = ttk.Frame(self.root)
#         self.auswahl_frame.pack(padx=20, pady=20)

#         self.zahnarzt_label = ttk.Label(self.auswahl_frame, text="Wählen Sie einen Zahnarzt:")
#         self.zahnarzt_label.grid(row=0, column=0, padx=10, pady=10)

#         self.zahnarzt_var = tk.StringVar()
#         self.zahnarzt_combobox = ttk.Combobox(self.auswahl_frame, textvariable=self.zahnarzt_var, values=list(self.zahnaerzte.keys()))
#         self.zahnarzt_combobox.grid(row=0, column=1, padx=10, pady=10)

#         self.zahnarzt_button = ttk.Button(self.auswahl_frame, text="Zahnarzt auswählen", command=self.anzeigen_terminkalender)
#         self.zahnarzt_button.grid(row=0, column=2, padx=10, pady=10)

#         # Variable pour stocker le rendez-vous sélectionné
#         self.selection_rendezvous = None

#     def generer_rendezvous_pour_annee(self):
#         # Générer des rendez-vous pour une année entière, toutes les demi-heures
#         debut_annee = datetime.datetime.now().replace(hour=8, minute=0, second=0, microsecond=0)
#         fin_annee = debut_annee + datetime.timedelta(days=365)
#         rendezvous = [debut_annee + datetime.timedelta(minutes=30 * i) for i in range(0, int((fin_annee - debut_annee).total_seconds() / 1800))]
#         return [rendezvous[i].strftime("%Y-%m-%d %H:%M") for i in range(len(rendezvous))]

#     def anzeigen_terminkalender(self):
#         ausgewaehlter_zahnarzt = self.zahnarzt_var.get()
#         if ausgewaehlter_zahnarzt:
#             # Erstellen Sie eine neue Top-Level-Fenster
#             terminkalender_fenster = tk.Toplevel(self.root)
#             terminkalender_fenster.title(f"Terminkalender für {ausgewaehlter_zahnarzt}")

#             terminkalender_treeview = ttk.Treeview(terminkalender_fenster, columns=('Tag', 'Datum', 'Uhrzeit'))
#             terminkalender_treeview.grid(row=0, column=0, padx=10, pady=10)

#             terminkalender_treeview.heading('Tag', text='Tag')
#             terminkalender_treeview.heading('Datum', text='Datum')
#             terminkalender_treeview.heading('Uhrzeit', text='Uhrzeit')

#             # Füllen Sie den Terminkalender mit Daten des ausgewählten Zahnarztes
#             termine = self.zahnaerzte[ausgewaehlter_zahnarzt]
#             for termin in termine:
#                 # Parse Datum und Uhrzeit
#                 datum, uhrzeit = termin.split()

#                 # Holen Sie den Wochentag
#                 tag = self.get_wochentag(datum)

#                 # Fügen Sie das Element zum Treeview hinzu
#                 terminkalender_treeview.insert('', 'end', values=(tag, datum, uhrzeit))

#             # Liez le double-clic à la fonction pour stocker le rendez-vous sélectionné
#             terminkalender_treeview.bind("<Double-1>", lambda event: self.selectionner_rendezvous(terminkalender_treeview))

#     def selectionner_rendezvous(self, terminkalender_treeview):
#         # Récupérer l'élément sélectionné
#         selection = terminkalender_treeview.selection()
        
#         if selection:
#             # Récupérer les détails du rendez-vous sélectionné
#             item = terminkalender_treeview.item(selection)
#             tag, date, heure = item['values']
#             rendezvous_details = f"Rendez-vous sélectionné : {tag}, {date} à {heure} pour le Zahnarzt {self.zahnarzt_var.get()}"
            
#             # Afficher les détails du rendez-vous dans le terminal
#             print(rendezvous_details)
            
#             # Stocker le rendez-vous dans la variable
#             self.selection_rendezvous = rendezvous_details

#     def get_wochentag(self, datum):
#         # Convertir la date en objet datetime
#         datetime_obj = datetime.datetime.strptime(datum, "%Y-%m-%d")
        
#         # Obtenir le nom du jour de la semaine (lundi à dimanche)
#         wochentag = datetime_obj.strftime("%A")
        
#         return wochentag

# if __name__ == "__main__":
#     root = tk.Tk()
#     app = ZahnarztTerminGUI(root)
#     root.mainloop()


import tkinter as tk
from tkinter import ttk
import datetime
import pandas as pd
from openpyxl import load_workbook

class ZahnarztTerminGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Zahnarzt Terminbuchung")

        self.zahnaerzte = {
            "Dr. Müller": [
                "2023-11-13 09:00", "2023-11-13 10:30", "2023-11-14 14:00",
                "2023-11-15 11:00", "2023-11-15 15:30", "2023-11-16 16:00"
            ],
            "Dr. Schmidt": [
                "2023-11-13 10:00", "2023-11-14 11:30", "2023-11-15 14:00",
                "2023-11-16 10:30", "2023-11-17 14:30", "2023-11-17 16:30"
            ]
        }

        self.create_widgets()

        self.selection_rendezvous = None

    def create_widgets(self):
        auswahl_frame = ttk.Frame(self.root)
        auswahl_frame.pack(padx=20, pady=20)

        ttk.Label(auswahl_frame, text="Wählen Sie einen Zahnarzt:").grid(row=0, column=0, padx=10, pady=10)

        self.zahnarzt_var = tk.StringVar()
        zahnarzt_combobox = ttk.Combobox(auswahl_frame, textvariable=self.zahnarzt_var, values=list(self.zahnaerzte.keys()))
        zahnarzt_combobox.grid(row=0, column=1, padx=10, pady=10)

        ttk.Button(auswahl_frame, text="Zahnarzt auswählen", command=self.show_appointments).grid(row=0, column=2, padx=10, pady=10)

    def show_appointments(self):
        selected_dentist = self.zahnarzt_var.get()
        if selected_dentist:
            appointments_window = tk.Toplevel(self.root)
            appointments_window.title(f"Terminkalender für {selected_dentist}")

            treeview = ttk.Treeview(appointments_window, columns=('Tag', 'Datum', 'Uhrzeit'))
            treeview.grid(row=0, column=0, padx=10, pady=10)

            for heading in ['Tag', 'Datum', 'Uhrzeit']:
                treeview.heading(heading, text=heading)

            appointments = self.zahnaerzte[selected_dentist]
            for appointment in appointments:
                day, date, time = self.parse_appointment(appointment)

                if 8 <= int(time.split(':')[0]) < 12 or 14 <= int(time.split(':')[0]) < 16:
                    treeview.insert('', 'end', values=(day, date, time))

            treeview.bind("<Double-1>", lambda event: self.select_appointment(treeview))

    def parse_appointment(self, appointment):
        date, time = appointment.split()
        day = self.get_weekday(date)
        return day, date, time

    def select_appointment(self, treeview):
        selection = treeview.selection()

        if selection:
            item = treeview.item(selection)
            day, date, time = item['values']
            appointment_details = f"Rendez-vous sélectionné : {day}, {date} à {time} pour le Zahnarzt {self.zahnarzt_var.get()}"
            print(appointment_details)
            self.selection_rendezvous = appointment_details

            # Stocker le rendez-vous dans la feuille 'privat' du fichier Excel
            self.save_appointment_to_excel(day, date, time)

    def save_appointment_to_excel(self, day, date, time):
        excel_file_path = 'rendezvous.xlsx'
        try:
            # Charger le classeur Excel existant
            wb = load_workbook(excel_file_path)

            # Sélectionner la feuille 'privat' (créer si elle n'existe pas)
            if 'Privat' not in wb.sheetnames:
                wb.create_sheet('Privat')

            ws = wb['Privat']

            # Ajouter une nouvelle ligne avec les détails du rendez-vous
            ws.append([self.zahnarzt_var.get(), day, date, time])

            # Enregistrer les modifications
            wb.save(excel_file_path)
            print(f"Rendez-vous enregistré dans {excel_file_path}, feuille 'Privat'")
        except FileNotFoundError:
            print("Le fichier Excel n'a pas été trouvé.")

    @staticmethod
    def get_weekday(date):
        datetime_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
        weekday = datetime_obj.strftime("%A")
        return weekday

if __name__ == "__main__":
    root = tk.Tk()
    app = ZahnarztTerminGUI(root)
    root.mainloop()
